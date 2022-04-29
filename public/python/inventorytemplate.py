from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from pymysql import*

wb = Workbook()

wss = wb.active
wsss = wb.create_sheet('Personnel')
ws = wb.create_sheet('Inventory')

ws.cell(row=1, column=3).value = "Article"
ws.cell(row=1, column=4).value = "Description"
ws.cell(row=1, column=5).value = "Date Acquired"
ws.cell(row=1, column=6).value = "Property Number"
ws.cell(row=1, column=7).value = "Unit of Measure"
ws.cell(row=1, column=8).value = "Unit Value"
ws.cell(row=1, column=9).value = "Total"
ws.cell(row=1, column=10).value = "Registered Employee?"
ws.cell(row=1, column=11).value = "Assigned to"
ws.cell(row=1, column=12).value = "Temp Name"
ws.cell(row=1, column=13).value = "Location"
ws.cell(row=1, column=14).value = "Remarks"

a1 = ws.cell(row=1, column=3)
a2 = ws.cell(row=1, column=4)
a3 = ws.cell(row=1, column=5)
a4 = ws.cell(row=1, column=6)
a5 = ws.cell(row=1, column=7)
a6 = ws.cell(row=1, column=8)
a7 = ws.cell(row=1, column=9)
a8 = ws.cell(row=1, column=10)
a9 = ws.cell(row=1, column=11)
a10 = ws.cell(row=1, column=12)
a11 = ws.cell(row=1, column=13)
a12 = ws.cell(row=1, column=14)

ft = Font(name='Arial', size=11, bold= True)
a1.font = ft
a2.font = ft
a3.font = ft
a4.font = ft
a5.font = ft
a6.font = ft
a7.font = ft
a8.font = ft
a9.font = ft
a10.font = ft
a11.font = ft
a12.font = ft

# connect the mysql with the python
con=connect(user="root",password="",host="localhost",database="inventorydb")
con1=connect(user="root",password="",host="localhost",database="personneldb")
# read the data
cursor = con.cursor()
cursor1 = con1.cursor()
cursor.execute("select id, article from article")
cursor1.execute("select SID,EMP_NO,CONCAT(NAME_F,' ',NAME_M,' ',NAME_L) as fullname from pds_personal_info")
data = cursor.fetchall()
data1 = cursor1.fetchall()

for row in data:
        wss.append(row)

for row in data1:
        wsss.append(row)
#You can change =$A:$A with a smaller range like =A1:A9
data_val = DataValidation(type="list",formula1='=Sheet!$B:$B',allow_blank=False) 
data_val1 = DataValidation(type="list",formula1='"YES,NO"',allow_blank=False)
data_val2 = DataValidation(type="list",formula1='=Personnel!$C:$C',allow_blank=False)

ws.add_data_validation(data_val)
ws.add_data_validation(data_val1)
ws.add_data_validation(data_val2)

#formula for article id
for i,row in enumerate(ws.iter_rows(min_col=1, max_col=1,min_row=1, max_row=20000)):
        for cell in row:
                cell.value = '=IFERROR(INDEX(Sheet!$A:$A,MATCH(Inventory!C{},Sheet!$B:$B,0))," ")'.format(i+1)
#formula for employee id
for i,row in enumerate(ws.iter_rows(min_col=2, max_col=2,min_row=1, max_row=20000)):
        for cell in row:
                cell.value = '=IFERROR(INDEX(Personnel!$A:$A,MATCH(Inventory!K{},Personnel!$C:$C,0))," ")'.format(i+1)
for col in ['A', 'B']:
    ws.column_dimensions[col].hidden= True
#If you go to the cell B1 you will find a drop down list with all the values from the column A
data_val.add(ws["C2"])
data_val1.add(ws["J2"]) 
data_val2.add(ws["K2"])
#data_val2.add(ws["A2"])
#data_val2.add(ws["B2"])

data_val.add('C2:C20000')
data_val1.add('J2:J20000')
data_val2.add('K2:K20000')
#data_val3.add('A2:A20000')
#data_val4.add('B2:B20000')

wss.sheet_state = 'hidden'
wsss.sheet_state = 'hidden'
wb.save('InventoryTemplate.csv')